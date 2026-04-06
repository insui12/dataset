"""수집 현황 자동 리포트 생성.

모든 팀(A/B/C/D)의 _state 폴더를 읽어 CSV + Excel 자동 생성.
cron으로 주기 실행 가능:
    */30 * * * * cd /home/selab/dataset && python3 scripts/generate_status_report.py
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_root / "src"))

ARTIFACTS = _root / "artifacts"
STATE_DIRS = {
    "B": ARTIFACTS / "json_downloads_round_robin_B" / "_state",
    "D": ARTIFACTS / "json_downloads_round_robin_D" / "_state",
    # A/C가 생기면 자동 감지
    "A": ARTIFACTS / "json_downloads_round_robin_A" / "_state",
    "C": ARTIFACTS / "json_downloads_round_robin_C" / "_state",
}

TEAM_NAMES = {"A": "A(규민)", "B": "B(서진)", "C": "C(혜린)", "D": "D(실습실)"}

# 예상 이슈 수 매핑 로드
def load_expected_issues():
    """bugzilla/jira/manifest에서 entry별 예상 이슈 수 로드."""
    expected = {}

    # Bugzilla
    bz_path = ARTIFACTS / "bugzilla_all_products.json"
    if bz_path.exists():
        with open(bz_path, encoding="utf-8") as f:
            for p in json.load(f):
                if isinstance(p, dict):
                    key = (p.get("instance", ""), p.get("entry_name", ""))
                    expected[key] = p.get("approx_issues", 0)

    # JIRA
    jira_path = ARTIFACTS / "apache_jira_projects.json"
    if jira_path.exists():
        with open(jira_path, encoding="utf-8") as f:
            for p in json.load(f):
                key = ("apache", p["key"].lower())
                expected[key] = p.get("issue_count", 0)

    # GitHub (PR 제외한 실제 issue 수 — B팀 수집 완료 결과 기준)
    # 기존 추정치는 issues+PRs 합산이라 과다 → 실측치로 교체
    gh_estimates = {
        ("github.com", "microsoft/vscode"): 238718,
        ("github.com", "llvm/llvm-project"): 94427,
        ("github.com", "nodejs/node"): 20877,
        ("github.com", "moby/moby"): 24149,
        # 아직 미수집 repos — 원래 추정치 유지 (PR 비율 약 50%로 가정)
        ("github.com", "rust-lang/rust"): 77000,
        ("github.com", "python/cpython"): 73000,
        ("github.com", "kubernetes/kubernetes"): 69000,
        ("github.com", "apache/airflow"): 32000,
    }
    expected.update(gh_estimates)

    # GitLab
    gl_estimates = {
        ("gitlab.com", "gitlab-org/gitlab"): 593744,
        ("gitlab.com", "gitlab-org/gitlab-runner"): 39322,
        ("gitlab.com", "gitlab-org/gitaly"): 7115,
        ("gitlab.com", "gitlab-org/omnibus-gitlab"): 9689,
    }
    expected.update(gl_estimates)

    # FreeBSD
    expected[("freebsd", "freebsd")] = 50727

    return expected


def read_states(state_dir: Path):
    """state 디렉토리에서 모든 state 파일 읽기."""
    if not state_dir.exists():
        return []
    results = []
    for sf in sorted(state_dir.iterdir()):
        if sf.suffix != ".json":
            continue
        try:
            with open(sf, encoding="utf-8") as f:
                st = json.load(f)
            results.append(st)
        except (json.JSONDecodeError, OSError):
            continue
    return results


def classify_source(family: str, instance: str) -> str:
    """출처 분류."""
    if family == "jira":
        return "JIRA"
    if family == "github":
        return "GitHub"
    if family == "gitlab":
        return "GitLab"
    if family == "bugzilla":
        name_map = {
            "mozilla": "Mozilla Bugzilla",
            "eclipse": "Eclipse Bugzilla",
            "freebsd": "FreeBSD Bugzilla",
            "gcc": "GCC Bugzilla",
            "kernel": "Kernel Bugzilla",
            "libreoffice": "LibreOffice Bugzilla",
        }
        return name_map.get(instance, f"{instance} Bugzilla")
    return family


def count_collected_files(base_dir: Path) -> dict:
    """실제 데이터 디렉토리에서 entry별 BASE 파일 수 집계.

    디렉토리 규칙: {FAMILY}_{INSTANCE}_{ENTRY} (uppercase, / and - → _)
    state 파일과 무관하게 실제 수집된 이슈 수를 측정 (race condition 영향 없음).
    """
    counts = {}
    if not base_dir.exists():
        return counts
    for entry_dir in base_dir.iterdir():
        if not entry_dir.is_dir() or entry_dir.name.startswith("_"):
            continue
        count = 0
        for root, dirs, files in os.walk(entry_dir):
            if os.path.basename(root) == "BASE":
                count += sum(1 for f in files if f.endswith(".json"))
        counts[entry_dir.name] = count
    return counts


def canonical_dir_name(family: str, instance: str, entry: str) -> str:
    """family/instance/entry → 디렉토리명 변환 (download_manifest_json.py 규칙)."""
    return f"{family}_{instance}_{entry}".replace("/", "_").replace("-", "_").upper()


def estimate_actual_issues(st: dict, expected_issues: int, file_counts: dict = None) -> int:
    """실제 수집된 이슈 수 반환.

    우선순위:
    1. 실제 BASE 파일 수 (file_counts 제공 시) — 가장 정확
    2. state 기반 추정 (fallback)
    """
    family = st.get("family", "")
    instance = st.get("instance", "")
    entry = st.get("entry", "")

    # 1순위: 실제 파일 수
    if file_counts is not None:
        dir_name = canonical_dir_name(family, instance, entry)
        if dir_name in file_counts:
            return file_counts[dir_name]

    # 2순위: state fallback
    completed = st.get("completed", False)
    pages = st.get("pages_completed", 0) or 0
    saved = st.get("issues_saved", 0) or 0

    if completed and expected_issues > 0:
        return expected_issues
    if family == "jira":
        return saved
    estimated = pages * 100
    if expected_issues > 0:
        return min(estimated, expected_issues)
    return estimated


def aggregate_team(states: list, expected: dict, file_counts: dict = None) -> dict:
    """팀 state를 출처별로 집계."""
    by_source = {}
    for st in states:
        family = st.get("family", "")
        instance = st.get("instance", "")
        source = classify_source(family, instance)
        completed = st.get("completed", False)

        entry_name = st.get("entry", "")
        # rename 역매핑
        if entry_name.startswith("jira_"):
            lookup_name = entry_name[5:]  # jira_incubator → incubator
        else:
            lookup_name = entry_name
        exp = expected.get((instance, lookup_name), 0) or expected.get(("apache", lookup_name), 0)

        actual_issues = estimate_actual_issues(st, exp, file_counts)

        if source not in by_source:
            by_source[source] = {
                "entries": 0, "expected_issues": 0,
                "collected_issues": 0, "completed_entries": 0,
            }
        by_source[source]["entries"] += 1
        by_source[source]["expected_issues"] += exp
        by_source[source]["collected_issues"] += actual_issues
        if completed:
            by_source[source]["completed_entries"] += 1

    return by_source


def get_d_machine_breakdown(state_dir: Path, expected: dict, file_counts: dict = None) -> list:
    """D팀 머신별 분류 (state에서 역추적 불가 → 배정 로직 재현)."""
    import yaml

    states = read_states(state_dir)
    if not states:
        return []

    # state를 entry명으로 인덱싱
    state_map = {}
    for st in states:
        entry = st.get("entry", "")
        state_map[entry] = st

    # 머신별 배정 재현
    jira_path = ARTIFACTS / "apache_jira_projects.json"
    with open(jira_path, encoding="utf-8") as f:
        projects = json.load(f)
    projects = sorted([p for p in projects if p["issue_count"] > 0],
                      key=lambda x: -x["issue_count"])
    rename = {"incubator": "jira_incubator", "testing": "jira_testing", "tools": "jira_tools"}

    ta_path = _root / "manifests" / "team_assignments.yaml"
    manifest_path = _root / "manifests" / "sample.manifest.yaml"
    c_mozilla = []
    if ta_path.exists() and manifest_path.exists():
        with open(ta_path, encoding="utf-8") as f:
            c_entries = yaml.safe_load(f)["teams"]["C"]
        with open(manifest_path, encoding="utf-8") as f:
            manifest = yaml.safe_load(f)
        mozilla_names = set()
        for fam in manifest["families"]:
            if fam["slug"] == "bugzilla":
                for inst in fam["instances"]:
                    if inst["name"] == "mozilla":
                        mozilla_names.update(e["name"] for e in inst["entries"])
        c_mozilla = [e for e in c_entries if e in mozilla_names]

    # 3/30 초기 수집 baseline: 각 entry에 대해 ~110 파일이 기본
    # 실제 작동: file_count > SEED_THRESHOLD (초기 수집분 초과)
    SEED_THRESHOLD = 110

    TOTAL = 41
    machines = []
    for m in range(1, TOTAL + 1):
        # JIRA entries
        my_entries = []
        for i, p in enumerate(projects):
            if (i % TOTAL) + 1 == m:
                name = rename.get(p["key"].lower(), p["key"].lower())
                my_entries.append(name)
        # Mozilla entries
        for i, name in enumerate(c_mozilla):
            if (i % TOTAL) + 1 == m:
                my_entries.append(name)

        total_expected = 0
        total_collected = 0
        entries_complete = 0  # 완료된 entries 수 (파일 수 >= 예상의 95%)
        entries_active = 0  # 초기 수집분 초과하여 작업된 entries
        entries_large = 0  # expected > 110인 entries (초기 수집으로 완료 불가)
        for entry in my_entries:
            exp = expected.get(("apache", entry), 0)
            if exp == 0:
                for inst in ["mozilla"]:
                    exp = expected.get((inst, entry), 0)
                    if exp: break
            total_expected += exp

            st = state_map.get(entry)
            # 파일 기반 실측 수집 이슈 수
            if st:
                actual = estimate_actual_issues(st, exp, file_counts)
            else:
                # state 없어도 파일로만 카운트
                if file_counts:
                    dir_name = f"jira_apache_{entry}".replace("/", "_").replace("-", "_").upper()
                    actual = file_counts.get(dir_name, 0)
                    if actual == 0:
                        dir_name = f"bugzilla_mozilla_{entry}".replace("/", "_").replace("-", "_").upper()
                        actual = file_counts.get(dir_name, 0)
                else:
                    actual = 0
            total_collected += actual

            # 완료 판정: 파일 수 >= 예상의 95%
            if exp > 0 and actual >= exp * 0.95:
                entries_complete += 1
            # 대규모 entry (초기 수집으로 완료 불가능한 entries)
            if exp > SEED_THRESHOLD:
                entries_large += 1
                # 실작동: file_count > 110 = post-deploy 작업 발생
                if actual > SEED_THRESHOLD:
                    entries_active += 1

        pct = (total_collected / total_expected * 100) if total_expected > 0 else 0

        # 상태 판정 (대규모 entries 기준)
        if entries_complete == len(my_entries) and len(my_entries) > 0:
            status = "완료"
        elif entries_large == 0:
            # 배정된 entries가 전부 소규모 (3/30에 이미 완료) - 매우 드물다
            status = "완료" if entries_complete == len(my_entries) else "진행중"
        elif entries_active == 0:
            status = "미작동"
        elif entries_active == entries_large:
            status = "진행중"
        else:
            status = f"일부({entries_active}/{entries_large})"

        all_completed = (entries_complete == len(my_entries) and len(my_entries) > 0)

        machines.append({
            "machine": m,
            "entries": len(my_entries),
            "expected": total_expected,
            "collected": total_collected,
            "progress": pct,
            "completed": all_completed and len(my_entries) > 0,
            "entries_active": entries_active,
            "status": status,
        })

    return machines


def generate_csv(rows: list, out_path: Path):
    """CSV 파일 생성."""
    header = "팀,출처,엔트리수,예상_전체이슈수,수집된_이슈수,수집_진행률\n"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(header)
        for r in rows:
            f.write(f"{r['team']},{r['source']},{r['entries']},"
                    f"{r['expected']:,},{r['collected']:,},{r['progress']:.1f}%\n")


def generate_xlsx(rows: list, d_machines: list, out_path: Path):
    """색상 코딩된 Excel 파일 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: 팀별 요약 ---
    ws = wb.active
    ws.title = "팀별 요약"

    headers = ["팀", "출처", "엔트리수", "예상 이슈수", "수집 이슈수", "진행률"]
    widths = [15, 22, 10, 15, 15, 12]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    white_fill = PatternFill("solid", fgColor="FFFFFF")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True, size=11)

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w

    for row_idx, r in enumerate(rows, 2):
        is_total = "합계" in r["team"]
        pct = r["progress"]

        vals = [r["team"], r["source"], r["entries"], r["expected"], r["collected"],
                f"{pct:.1f}%"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if col >= 3:
                cell.alignment = Alignment(horizontal="right")

            if is_total:
                cell.fill = gray_fill
                cell.font = bold
            elif pct >= 100:
                cell.fill = white_fill
            elif pct >= 20:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    ws.freeze_panes = "A2"

    # --- Sheet 2: D팀 머신별 ---
    if d_machines:
        ws2 = wb.create_sheet("D팀 머신별")
        headers2 = ["머신", "엔트리수", "예상 이슈수", "수집 이슈수", "진행률", "작동상태"]
        widths2 = [10, 10, 15, 15, 12, 18]

        for i, (h, w) in enumerate(zip(headers2, widths2), 1):
            cell = ws2.cell(row=1, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws2.column_dimensions[chr(64 + i)].width = w

        for row_idx, m in enumerate(d_machines, 2):
            pct = m["progress"]
            status = m.get("status", "진행중")
            is_inactive = status == "미작동"
            vals = [f"M{m['machine']}", m["entries"], m["expected"],
                    m["collected"], f"{pct:.1f}%", status]

            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.border = border
                if col >= 2:
                    cell.alignment = Alignment(horizontal="right")
                # 미작동은 빨강, 완료는 흰색, 나머지는 초록
                if is_inactive:
                    cell.fill = red_fill
                elif m["completed"]:
                    cell.fill = white_fill
                else:
                    cell.fill = green_fill

        # 합계 행
        total_row = len(d_machines) + 2
        total_exp = sum(m["expected"] for m in d_machines)
        total_col = sum(m["collected"] for m in d_machines)
        total_pct = (total_col / total_exp * 100) if total_exp > 0 else 0
        done_count = sum(1 for m in d_machines if m["completed"])
        inactive_count = sum(1 for m in d_machines if m.get("status") == "미작동")
        vals = ["합계", len(d_machines), total_exp, total_col,
                f"{total_pct:.1f}%", f"완료{done_count}/미작동{inactive_count}"]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=total_row, column=col, value=val)
            cell.border = border
            cell.fill = gray_fill
            cell.font = bold
            if col >= 2:
                cell.alignment = Alignment(horizontal="right")

        ws2.freeze_panes = "A2"

    wb.save(out_path)


SPREADSHEET_KEY = "1tDZW4JyINgOmsA5rZd9kTUatf4Rt8ZIQMZ92TT-qWIc"
TOKEN_PATH = Path.home() / ".config" / "gspread" / "authorized_user.json"


def upload_to_google_sheets(rows: list, d_machines: list):
    """Google Sheets에 현황 자동 업로드."""
    if not TOKEN_PATH.exists():
        print("  [SKIP] Google Sheets: 토큰 없음")
        return

    try:
        import gspread
        from google.oauth2.credentials import Credentials

        with open(TOKEN_PATH) as f:
            token = json.load(f)

        creds = Credentials(
            token=token["token"],
            refresh_token=token["refresh_token"],
            token_uri=token["token_uri"],
            client_id=token["client_id"],
            client_secret=token["client_secret"],
            scopes=token["scopes"],
        )
        # 만료 시 갱신
        if creds.expired or not creds.valid:
            from google.auth.transport.requests import Request
            creds.refresh(Request())
            token["token"] = creds.token
            with open(TOKEN_PATH, "w") as f:
                json.dump(token, f, indent=2)

        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_KEY)

        # --- Sheet 1: 팀별 요약 ---
        try:
            ws = sh.worksheet("팀별 요약")
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet("팀별 요약", rows=len(rows) + 2, cols=7)

        header = ["팀", "출처", "엔트리수", "예상 이슈수", "수집 이슈수", "진행률", "업데이트"]
        data = [header]
        now_str = datetime.now().strftime("%m/%d %H:%M")
        for r in rows:
            data.append([
                r["team"], r["source"], r["entries"],
                f"{r['expected']:,}", f"{r['collected']:,}",
                f"{r['progress']:.1f}%", now_str,
            ])
        ws.clear()
        ws.update(data, "A1")

        # 서식 일괄 적용 (batch_format: API 1회)
        fmt_list = [
            {"range": "A1:G1", "format": {
                "backgroundColor": {"red": 0.18, "green": 0.33, "blue": 0.59},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "horizontalAlignment": "CENTER",
            }},
        ]
        for i, r in enumerate(rows, 2):
            is_total = "합계" in r["team"]
            pct = r["progress"]
            if is_total:
                bg = {"red": 0.85, "green": 0.85, "blue": 0.85}
                fmt_list.append({"range": f"A{i}:G{i}", "format": {"backgroundColor": bg, "textFormat": {"bold": True}}})
            else:
                if pct >= 100:
                    bg = {"red": 1, "green": 1, "blue": 1}
                elif pct >= 20:
                    bg = {"red": 0.78, "green": 0.94, "blue": 0.81}
                else:
                    bg = {"red": 1, "green": 0.78, "blue": 0.81}
                fmt_list.append({"range": f"A{i}:G{i}", "format": {"backgroundColor": bg}})
        ws.batch_format(fmt_list)
        ws.freeze(rows=1)

        # --- Sheet 2: D팀 머신별 ---
        if d_machines:
            try:
                ws2 = sh.worksheet("D팀 머신별")
                ws2.clear()
            except gspread.exceptions.WorksheetNotFound:
                ws2 = sh.add_worksheet("D팀 머신별", rows=len(d_machines) + 3, cols=7)

            header2 = ["머신", "엔트리수", "예상 이슈수", "수집 이슈수", "진행률", "작동상태", "업데이트"]
            data2 = [header2]
            for m in d_machines:
                data2.append([
                    f"M{m['machine']}", m["entries"],
                    f"{m['expected']:,}", f"{m['collected']:,}",
                    f"{m['progress']:.1f}%",
                    m.get("status", "진행중"),
                    now_str,
                ])
            total_exp = sum(m["expected"] for m in d_machines)
            total_col = sum(m["collected"] for m in d_machines)
            total_pct = (total_col / total_exp * 100) if total_exp > 0 else 0
            done_count = sum(1 for m in d_machines if m["completed"])
            inactive_count = sum(1 for m in d_machines if m.get("status") == "미작동")
            data2.append([
                "합계", len(d_machines), f"{total_exp:,}", f"{total_col:,}",
                f"{total_pct:.1f}%",
                f"완료{done_count}/미작동{inactive_count}",
                now_str,
            ])
            ws2.update(data2, "A1")

            fmt_list2 = [
                {"range": "A1:G1", "format": {
                    "backgroundColor": {"red": 0.18, "green": 0.33, "blue": 0.59},
                    "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    "horizontalAlignment": "CENTER",
                }},
            ]
            for i, m in enumerate(d_machines, 2):
                status = m.get("status", "")
                if status == "미작동":
                    bg = {"red": 1, "green": 0.78, "blue": 0.81}  # 빨강
                elif m["completed"]:
                    bg = {"red": 1, "green": 1, "blue": 1}  # 흰색
                else:
                    bg = {"red": 0.78, "green": 0.94, "blue": 0.81}  # 초록
                fmt_list2.append({"range": f"A{i}:G{i}", "format": {"backgroundColor": bg}})
            last_row = len(d_machines) + 2
            fmt_list2.append({"range": f"A{last_row}:G{last_row}", "format": {
                "backgroundColor": {"red": 0.85, "green": 0.85, "blue": 0.85},
                "textFormat": {"bold": True},
            }})
            ws2.batch_format(fmt_list2)
            ws2.freeze(rows=1)

        print("  Google Sheets 업데이트 완료")

    except Exception as e:
        print(f"  [ERROR] Google Sheets: {e}")


def main():
    expected = load_expected_issues()
    rows = []

    # 각 팀의 실제 데이터 파일 수 카운트 (한 번만)
    file_counts_by_team = {}
    for team in ["A", "B", "C", "D"]:
        team_dir = ARTIFACTS / f"json_downloads_round_robin_{team}"
        file_counts_by_team[team] = count_collected_files(team_dir)

    for team in ["A", "B", "C"]:
        state_dir = STATE_DIRS[team]
        states = read_states(state_dir)
        if not states:
            # 데이터 없으면 스킵
            continue
        by_source = aggregate_team(states, expected, file_counts_by_team[team])
        team_total_exp = 0
        team_total_col = 0

        for source in sorted(by_source.keys()):
            d = by_source[source]
            pct = (d["collected_issues"] / d["expected_issues"] * 100) if d["expected_issues"] > 0 else 0
            rows.append({
                "team": TEAM_NAMES[team], "source": source,
                "entries": d["entries"],
                "expected": d["expected_issues"],
                "collected": d["collected_issues"],
                "progress": pct,
            })
            team_total_exp += d["expected_issues"]
            team_total_col += d["collected_issues"]

        # 팀 합계
        pct = (team_total_col / team_total_exp * 100) if team_total_exp > 0 else 0
        rows.append({
            "team": f"{TEAM_NAMES[team]} 합계", "source": "",
            "entries": sum(d["entries"] for d in by_source.values()),
            "expected": team_total_exp,
            "collected": team_total_col,
            "progress": pct,
        })

    # D팀
    d_machines = []
    state_dir = STATE_DIRS["D"]
    if state_dir.exists():
        try:
            import yaml
            d_machines = get_d_machine_breakdown(state_dir, expected, file_counts_by_team["D"])
        except ImportError:
            d_machines = []

        states = read_states(state_dir)
        if states:
            by_source = aggregate_team(states, expected, file_counts_by_team["D"])
            team_total_exp = 0
            team_total_col = 0

            for source in sorted(by_source.keys()):
                d = by_source[source]
                pct = (d["collected_issues"] / d["expected_issues"] * 100) if d["expected_issues"] > 0 else 0
                rows.append({
                    "team": TEAM_NAMES["D"], "source": source,
                    "entries": d["entries"],
                    "expected": d["expected_issues"],
                    "collected": d["collected_issues"],
                    "progress": pct,
                })
                team_total_exp += d["expected_issues"]
                team_total_col += d["collected_issues"]

            pct = (team_total_col / team_total_exp * 100) if team_total_exp > 0 else 0
            rows.append({
                "team": f"{TEAM_NAMES['D']} 합계", "source": "",
                "entries": sum(d["entries"] for d in by_source.values()),
                "expected": team_total_exp,
                "collected": team_total_col,
                "progress": pct,
            })

    # 총합계
    all_exp = sum(r["expected"] for r in rows if "합계" in r["team"])
    all_col = sum(r["collected"] for r in rows if "합계" in r["team"])
    all_ent = sum(r["entries"] for r in rows if "합계" in r["team"])
    pct = (all_col / all_exp * 100) if all_exp > 0 else 0
    rows.append({
        "team": "총합계", "source": "",
        "entries": all_ent,
        "expected": all_exp,
        "collected": all_col,
        "progress": pct,
    })

    # 출력
    csv_path = ARTIFACTS / "collection_summary_by_team.csv"
    xlsx_path = ARTIFACTS / "버그리포트_수집_현황.xlsx"

    generate_csv(rows, csv_path)
    generate_xlsx(rows, d_machines, xlsx_path)

    # Google Sheets 업데이트
    upload_to_google_sheets(rows, d_machines)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{now}] 리포트 생성 완료")
    print(f"  CSV:  {csv_path}")
    print(f"  XLSX: {xlsx_path}")

    # 간단 요약 출력
    for r in rows:
        if "합계" in r["team"] or r["team"] == "총합계":
            print(f"  {r['team']}: {r['collected']:,} / {r['expected']:,} ({r['progress']:.1f}%)")


if __name__ == "__main__":
    main()
